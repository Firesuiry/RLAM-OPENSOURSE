import json
import numpy as np
import pickle
import re
import openpyxl
def write_xlsx(titles, datas, file_path, add=False):
    add = add and os.path.exists(file_path)
    if add:
        book = openpyxl.load_workbook(file_path)
        sheet = book['files']
    else:
        book = openpyxl.Workbook()
        sheet = book.create_sheet('files', 0)

    col = 1
    for title in titles:
        sheet.cell(row=1, column=col).value = title
        col += 1

    row = 2
    have_add_list = []
    if add:
        while True:
            d = sheet.cell(row=row, column=1).value
            if d:
                have_add_list.append(d)
                row += 1
            else:
                break

    for row_data in datas:
        if row_data[0] not in have_add_list:
            col = 1
            for data in row_data:
                sheet.cell(row=row, column=col).value = data
                col += 1
            row += 1
    book.save(file_path)

re_pattern = re.compile('(.*)-separate_train(.*)-group(.*)-dim(.*)', re.M|re.I)
with open('result.pickle','rb') as f:
    result = pickle.loads(f.read())
evaluate_result = result['result']
train_average_ranks = []
origin_average_ranks = []
for res in evaluate_result:
    train_average_ranks.append(res['result'][0]['train_average_rank'])
    origin_average_ranks.append(res['result'][0]['origin_average_rank'])
print(f'train_average_ranks:{np.mean(train_average_ranks)}')
print(f'train_average_ranks:{np.mean(origin_average_ranks)}')
def get_attr(type_str):
    print(type_str)
    searchObj  = re_pattern.search(type_str)
    data = {
        'name':searchObj.group(1),
        'separate_train':searchObj.group(2),
        'group':searchObj.group(3),
        'dim':searchObj.group(4),
    }
    return data

title = ['name','separate_train','group','dim','after_train','origin','improve']
final_data = []
def append_data(result):
    type_data = get_attr(result['type'])
    data = []
    for k,v in type_data.items(): # name  CLPSO
        data.append(v)
    data.append(result['result'][0]['train_average_rank'])
    data.append(result['result'][0]['origin_average_rank'])
    train_rank = result['result'][0]['train_average_rank']
    origin_rank = result['result'][0]['origin_average_rank']
    improve = (origin_rank-train_rank)/origin_rank*100
    data.append(improve)
    final_data.append(data)
for res in evaluate_result:
    append_data(res)

write_xlsx(title,final_data,'data.xlsx')



